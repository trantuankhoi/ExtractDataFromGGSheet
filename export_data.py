import os
import shutil
import requests
from tqdm import tqdm
from loguru import logger

import openpyxl

from config import IMAGE_DIR, GT_NAME
from config import CREDENTIAL_PATH, SCOPES, SHEET_DIR, FOLDER_IDS
from config import COLUMN_INSTANCE_ID, COLUMN_IMG_ID, COLUMN_GT_ID, CONFUSE_COLUMN_ID
from gg_drive_api import DriveAPI


class Exporter:
    def __init__(self):
        self.service = DriveAPI(cre_path=CREDENTIAL_PATH, scopes=SCOPES)
        self.folder_ids = FOLDER_IDS
        self.gt_content = []

    def download_sheets(self):
        for folder_id in FOLDER_IDS:
            items = self.service.get_files_in_folder(folder_id=folder_id)
            for item in items:
                file_id = item['id']
                file_name = item['name']

                if "text" in file_name:
                    file_path = os.path.join(SHEET_DIR, file_name + ".xlsx")
                    self.service.download_file_by_id(file_id=file_id, file_name=file_name, des_path=file_path)    

    def download_image(self, img_url, img_path):
        status = False

        while not status:
            try:
                response = requests.get(img_url, stream=True)
                with open(img_path, 'wb') as f:
                    shutil.copyfileobj(response.raw, f)
                status = True
                del response
            except:
                logger.warning(f"Trying download image: {img_url}")

    def extract(self, sheet_path):
        logger.info(f"Extracting file {os.path.basename(sheet_path)}")
        worksheet = openpyxl.load_workbook(sheet_path)

        for sheet_name in worksheet.sheetnames:
            sheet = worksheet[sheet_name]

            pbar = tqdm(total=sheet.max_row - 1, ncols=70, desc=sheet_name)
            for idx, row in enumerate(sheet.rows):
                if idx == 0 or row[COLUMN_IMG_ID].value is None:
                    continue
                
                confused = str(row[CONFUSE_COLUMN_ID].value).lower()
                assert confused in ["true", "false", "x", "", "none"] or confused is None, "Invalid sheet's format"
                # if confused not in ["true", "false", "x", "", "none"]:
                #     breakpoint()
                if confused in ['true', 'x']:
                    continue

                img_url = row[COLUMN_IMG_ID].value.split(",")[0].replace("=IMAGE(", "").replace('"', '')
                img_name = img_url.split("/")[-1]
                img_path = os.path.join(IMAGE_DIR, img_name)
                self.download_image(img_url=img_url, img_path=img_path)

                gt = row[COLUMN_GT_ID].value

                self.gt_content.append(f"{img_name}\t{gt}\n")
                pbar.update()

    def run(self):
        # Download sheets
        logger.info("Downloading sheets")
        self.download_sheets()

        # Extract sheet
        logger.info("Extracting sheet")
        for sheet in os.listdir(SHEET_DIR):
            if ".xlsx" not in sheet:
                continue
            sheet_path = os.path.join(SHEET_DIR, sheet)
            self.extract(sheet_path=sheet_path)

        # Save ground truth    
        gt_path = os.path.join(IMAGE_DIR, GT_NAME)
        with open(gt_path, "w") as f:
            f.writelines(self.gt_content)



exporter = Exporter()
exporter.run()
